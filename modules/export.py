"""
Module d'export Excel pour DocScan Pro.

Génère des fichiers Excel organisés : index général + feuilles par Client × Type.
"""

import io
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.config import TYPE_COLORS, TYPE_CONFIG


def safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Nettoie un nom pour l'utiliser comme nom de feuille Excel.

    Excel interdit les caractères : \\ / * ? : [ ]
    Limite à 31 caractères maximum.

    Args:
        name: Nom brut de la feuille.
        max_len: Longueur maximale (défaut 31, limite Excel).

    Returns:
        Nom de feuille valide pour Excel.
    """
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:max_len]


def format_worksheet(ws, header_color: str = "1e3a5f") -> None:
    """Applique un formatage professionnel à une feuille Excel.

    - En-têtes : fond coloré, texte blanc gras centré
    - Données : alternance de lignes, bordures fines
    - Colonnes : largeur auto-ajustée (min 10, max 45)
    - Première ligne figée

    Args:
        ws: Feuille openpyxl (Worksheet).
        header_color: Couleur de fond des en-têtes en hex sans # (défaut bleu marine).
    """
    header_fill = PatternFill(
        start_color=header_color, end_color=header_color, fill_type="solid"
    )
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell_font = Font(name="Calibri", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill(
        start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"
    )

    # Formater les en-têtes
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Formater les données
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-ajuster la largeur des colonnes
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        adjusted = min(max(max_length + 3, 10), 45)
        ws.column_dimensions[col_letter].width = adjusted

    # Figer la première ligne
    ws.freeze_panes = "A2"


def build_organized_excel(history: list[dict]) -> bytes:
    """Génère un fichier Excel multi-feuilles organisé par Client × Type.

    Structure :
    - Feuille "Index général" : tous les documents
    - Une feuille par combinaison Client × Type : résumé
    - Une feuille par combinaison Client × Type DET : lignes détaillées

    Args:
        history: Liste des documents traités (session_state.history).

    Returns:
        Contenu binaire du fichier Excel (.xlsx).
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # 1. Index général
        all_flats = []
        for h in history:
            row = h["flat"].copy()
            row["Fichier source"] = h["filename"]
            row["Date extraction"] = h["timestamp"]
            all_flats.append(row)

        pd.DataFrame(all_flats).to_excel(
            writer, sheet_name="Index général", index=False
        )
        format_worksheet(writer.sheets["Index général"], "1e3a5f")

        # 2. Par client → par type
        by_client: dict = defaultdict(list)
        for h in history:
            by_client[h["client"]].append(h)

        for client_name, client_docs in sorted(by_client.items()):
            by_type: dict = defaultdict(list)
            for doc in client_docs:
                by_type[doc["type"]].append(doc)

            for type_key, type_docs in sorted(by_type.items()):
                type_label = TYPE_CONFIG.get(type_key, {}).get("label", type_key)
                sheet_name = safe_sheet_name(f"{client_name} - {type_label}")
                color = TYPE_COLORS.get(type_key, "64748b")

                # Feuille résumé
                rows = []
                for doc in type_docs:
                    row = doc["flat"].copy()
                    row["Fichier source"] = doc["filename"]
                    rows.append(row)

                pd.DataFrame(rows).to_excel(
                    writer, sheet_name=sheet_name, index=False
                )
                format_worksheet(writer.sheets[sheet_name], color)

                # Feuille lignes détaillées
                all_lines = []
                for doc in type_docs:
                    if doc["lines_df"] is not None and not doc["lines_df"].empty:
                        ldf = doc["lines_df"].copy()
                        ldf.insert(0, "N° Document", doc["flat"].get("N° Document", ""))
                        ldf.insert(0, "Fichier", doc["filename"])
                        all_lines.append(ldf)

                if all_lines:
                    combined = pd.concat(all_lines, ignore_index=True)
                    lines_sheet = safe_sheet_name(f"{client_name} - {type_label} DET")
                    combined.to_excel(writer, sheet_name=lines_sheet, index=False)
                    format_worksheet(writer.sheets[lines_sheet], color)

    return output.getvalue()


def build_single_excel(flat: dict, lines_df: pd.DataFrame | None) -> bytes:
    """Génère un fichier Excel pour un document individuel.

    Structure :
    - Feuille "Résumé" : champs clés en format clé/valeur
    - Feuille "Lignes détaillées" : lignes de détail (si présentes)

    Args:
        flat: Dictionnaire plat des métadonnées du document.
        lines_df: DataFrame des lignes de détail, ou None.

    Returns:
        Contenu binaire du fichier Excel (.xlsx).
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Feuille résumé transposée pour une meilleure lisibilité
        summary_data = {"Champ": list(flat.keys()), "Valeur": list(flat.values())}
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Résumé", index=False)
        format_worksheet(writer.sheets["Résumé"], "1e3a5f")
        writer.sheets["Résumé"].column_dimensions["B"].width = 50

        if lines_df is not None and not lines_df.empty:
            lines_df.to_excel(writer, sheet_name="Lignes détaillées", index=False)
            format_worksheet(writer.sheets["Lignes détaillées"], "2563eb")

    return output.getvalue()
